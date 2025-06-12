import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os

# --- Configuration: Define the cell mappings ---
CELL_MAPPINGS = [
    ('Lening', 'C3', 'Gegevens', 'C3'), ('Lening', 'C4', 'Gegevens', 'C4'),
    ('Lening', 'C5', 'Gegevens', 'C5'), ('Lening', 'C6', 'Gegevens', 'C6'),
    ('Lening', 'C7', 'Gegevens', 'C7'), ('Lening', 'C8', 'Gegevens', 'C8'),
    ('Lening', 'C9', 'Gegevens', 'C9'), ('Lening', 'C10', 'Gegevens', 'C10'),
    ('Lening', 'C11', 'Gegevens', 'C11'), ('Lening', 'C12', 'Gegevens', 'C12'),
    ('Lening', 'C13', 'Gegevens', 'C13'), ('Lening', 'C14', 'Gegevens', 'C14'),
    ('Lening', 'C15', 'Gegevens', 'C15'), ('Lening', 'C17', 'Gegevens', 'C17'),
    ('Lening', 'C18', 'Gegevens', 'C18'), ('Lening', 'C19', 'Gegevens', 'C19'),
    ('Lening', 'C20', 'Gegevens', 'C20'), ('Lening', 'C21', 'Gegevens', 'C21'),
    ('Lening', 'C22', 'Gegevens', 'C22'), ('Lening', 'C24', 'Gegevens', 'C24'),
    ('Lening', 'C25', 'Gegevens', 'C25'), ('Lening', 'C26', 'Gegevens', 'C26'),
    ('Lening', 'C27', 'Gegevens', 'C27'), ('Lening', 'C28', 'Gegevens', 'C28'),
    ('Lening', 'C30', 'Gegevens', 'C30'), ('Lening', 'C31', 'Gegevens', 'C31'),
    ('Lening', 'C32', 'Gegevens', 'C32'), ('Lening', 'C33', 'Gegevens', 'C33'),
    ('Lening', 'C36', 'Gegevens', 'C36'), ('Lening', 'C37', 'Gegevens', 'C37'),
    ('Lening', 'C38', 'Gegevens', 'C38'), ('Lening', 'C39', 'Gegevens', 'C39'),
    ('Lening', 'C40', 'Gegevens', 'C40'), ('Lening', 'C51', 'Gegevens', 'C51'),
    ('Lening', 'C52', 'Gegevens', 'C52'), ('Lening', 'C53', 'Gegevens', 'C53'),
    ('Lening', 'C55', 'Gegevens', 'C55'), ('Lening', 'C56', 'Gegevens', 'C56'),
    ('Lening', 'C57', 'Gegevens', 'C57'), ('Lening', 'C58', 'Gegevens', 'C58'),
    ('Lening', 'E57', 'Gegevens', 'E57'), ('Lening', 'C60', 'Gegevens', 'C60'),
    ('Lening', 'C61', 'Gegevens', 'C61'), ('Lening', 'C62', 'Gegevens', 'C62'),
    ('Lening', 'C63', 'Gegevens', 'C63'), ('Lening', 'E60', 'Gegevens', 'E60'),
    ('Lening', 'E61', 'Gegevens', 'E61'), ('Lening', 'E63', 'Gegevens', 'E63'),
    ('Lening', 'C66', 'Gegevens', 'C66'), ('Lening', 'C67', 'Gegevens', 'C67'),
    ('Lening', 'C68', 'Gegevens', 'C68'), ('Lening', 'D66', 'Gegevens', 'D66'),
    ('Lening', 'D67', 'Gegevens', 'D67'), ('Lening', 'D68', 'Gegevens', 'D68'),
    ('Lening', 'C70', 'Gegevens', 'C70'), ('Lening', 'D70', 'Gegevens', 'D70'),
    ('Lening', 'F70', 'Gegevens', 'F70'), ('Lening', 'G70', 'Gegevens', 'G70'),
    ('Lening', 'C72', 'Gegevens', 'C72'), ('Lening', 'C73', 'Gegevens', 'C73'),
    ('Lening', 'C74', 'Gegevens', 'C74'), ('Lening', 'C75', 'Gegevens', 'C75'),
    ('Lening', 'C76', 'Gegevens', 'C76'), ('Lening', 'C77', 'Gegevens', 'C77'),
    ('Lening', 'C78', 'Gegevens', 'C78'), ('Lening', 'C79', 'Gegevens', 'C79'),
    ('Lening', 'C80', 'Gegevens', 'C80'), ('Lening', 'E73', 'Gegevens', 'E73'),
    ('Lening', 'E74', 'Gegevens', 'E74'), ('Lening', 'E75', 'Gegevens', 'E75'),
    ('Lening', 'E76', 'Gegevens', 'E76'), ('Lening', 'E77', 'Gegevens', 'E77'),
    ('Lening', 'E78', 'Gegevens', 'E78'), ('Lening', 'E79', 'Gegevens', 'E79'),
    ('Lening', 'E80', 'Gegevens', 'E80'), ('Lening', 'C82', 'Gegevens', 'C82'),
    ('Lening', 'C83', 'Gegevens', 'C83'), ('Lening', 'C85', 'Gegevens', 'C85'),
    ('Lening', 'C86', 'Gegevens', 'C86'), ('Lening', 'C87', 'Gegevens', 'C87'),
    ('Lening', 'D85', 'Gegevens', 'D85'), ('Lening', 'D86', 'Gegevens', 'D86'),
    ('Lening', 'D87', 'Gegevens', 'D87'), ('Lening', 'D89', 'Gegevens', 'D89'),
    ('Lening', 'E89', 'Gegevens', 'E89'), ('Lening', 'F89', 'Gegevens', 'F89'),
    ('Lening', 'C91', 'Gegevens', 'C91'), ('Lening', 'C94', 'Gegevens', 'C94'),
    ('Lening', 'C95', 'Gegevens', 'C95'),
]

# --- Core Logic for Data Transfer ---
def transfer_data_to_template(form_file_path, template_file_path, output_file_path):
    """Transfers data from the form to the template file using openpyxl."""
    try:
        wb_form = openpyxl.load_workbook(form_file_path, data_only=True)
        sheet_lening = wb_form['Lening']

        wb_template = openpyxl.load_workbook(template_file_path)
        sheet_gegevens = wb_template['Gegevens']

        for src_sheet_name, src_cell_addr, dest_sheet_name, dest_cell_addr in CELL_MAPPINGS:
            if src_sheet_name == 'Lening' and dest_sheet_name == 'Gegevens':
                value_to_transfer = sheet_lening[src_cell_addr].value
                sheet_gegevens[dest_cell_addr].value = value_to_transfer
        
        wb_template.save(output_file_path)
        return True, "Data successfully transferred."
    except FileNotFoundError:
        return False, "Error: One or both files not found. Please check paths."
    except KeyError as e:
        return False, f"Error: Sheet '{e.args[0]}' not found in one of the workbooks."
    except Exception as e:
        return False, f"An unexpected error occurred during data transfer: {str(e)}"

# --- REVISED: Function to clean sheets and return the dossier name ---
def process_and_clean_sheets(workbook_path):
    """
    Cleans DATA and FACTUUR sheets and returns the dossier name from 'Gegevens'!C3.
    """
    try:
        # --- Configuration ---
        GEGEVENS_SHEET = 'Gegevens'
        DOSSIER_CELL = 'C3'
        MAX_PERIODEN_CELL = 'C52'
        INTEREST_PAYMENT_CELL = 'C57'
        
        DATA_SHEET = 'DATA'
        FACTUUR_SHEET = 'FACTUUR'

        wb = openpyxl.load_workbook(workbook_path)

        # 1. Get operational values from 'Gegevens' sheet
        if GEGEVENS_SHEET not in wb.sheetnames:
            return False, f"Error: Sheet '{GEGEVENS_SHEET}' not found.", None
        gegevens_sheet = wb[GEGEVENS_SHEET]
        
        dossier_name = gegevens_sheet[DOSSIER_CELL].value or "Onbekend Dossier"
        
        max_perioden = gegevens_sheet[MAX_PERIODEN_CELL].value
        if not isinstance(max_perioden, (int, float)) or max_perioden <= 0:
            return False, f"Error: Loan term in {MAX_PERIODEN_CELL} ('{max_perioden}') is not a valid positive number.", None
        max_perioden = int(max_perioden)

        interest_payment_type = gegevens_sheet[INTEREST_PAYMENT_CELL].value
        if interest_payment_type == 'Maandelijks':
            data_start_row = 16
        else:
            data_start_row = 15

        rows_deleted_count = 0

        # 2. Clean the 'DATA' sheet
        if DATA_SHEET not in wb.sheetnames:
            return False, f"Error: Sheet '{DATA_SHEET}' not found.", None

        data_sheet = wb[DATA_SHEET]
        start_row_for_deletion = data_sheet.max_row - 1
        
        for row_num in range(start_row_for_deletion, data_start_row - 1, -1):
            calculated_period = row_num - (data_start_row - 1)
            if calculated_period > max_perioden:
                data_sheet.delete_rows(row_num)
                rows_deleted_count += 1

        # 3. Clean the 'FACTUUR' sheet
        if rows_deleted_count > 0:
            if FACTUUR_SHEET in wb.sheetnames:
                factuur_sheet = wb[FACTUUR_SHEET]
                for _ in range(rows_deleted_count):
                    if factuur_sheet.max_row > 0:
                        factuur_sheet.delete_rows(factuur_sheet.max_row)
                    else:
                        break
            else:
                wb.save(workbook_path)
                return True, (f"Successfully cleaned '{DATA_SHEET}'. "
                               f"Warning: Sheet '{FACTUUR_SHEET}' not found, so it was not cleaned."), dossier_name

        wb.save(workbook_path)
        message = (f"Successfully cleaned sheets. Removed {rows_deleted_count} rows "
                   f"from '{DATA_SHEET}' and '{FACTUUR_SHEET}'.")
        return True, message, dossier_name

    except Exception as e:
        return False, f"An error occurred while cleaning the sheets: {str(e)}", None

# --- New Helper Function to Sanitize Filenames ---
def sanitize_filename(name):
    """Removes characters that are invalid in Windows filenames."""
    return "".join(c for c in name if c not in '<>:"/\\|?*')

# --- GUI Implementation ---
class ExcelTransferApp:
    def __init__(self, master):
        self.master = master
        master.title("Excel Data Transfer Tool")
        master.geometry("850x250")
        self.template_path = tk.StringVar()
        self.form_path = tk.StringVar()
        self.status_message = tk.StringVar()
        self.status_message.set("Please select your files and click 'Transfer Data'.")

        tk.Label(master, text="Template file:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        tk.Entry(master, textvariable=self.template_path, width=70, state="readonly").grid(row=0, column=1, padx=5, pady=10)
        tk.Button(master, text="Browse...", command=self.browse_template).grid(row=0, column=2, padx=10, pady=10)
        tk.Button(master, text="Clear", command=self.clear_template_path).grid(row=0, column=3, padx=10, pady=10)

        tk.Label(master, text="Filled-in Form:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        tk.Entry(master, textvariable=self.form_path, width=70, state="readonly").grid(row=1, column=1, padx=5, pady=10)
        tk.Button(master, text="Browse...", command=self.browse_form).grid(row=1, column=2, padx=10, pady=10)
        tk.Button(master, text="Clear", command=self.clear_form_path).grid(row=1, column=3, padx=10, pady=10)
        
        tk.Button(master, text="Transfer Data & Clean", command=self.run_transfer, height=2, width=25, bg="#4CAF50", fg="white").grid(row=2, column=0, columnspan=4, pady=20)
        tk.Label(master, textvariable=self.status_message, fg="blue", wraplength=800, justify="left").grid(row=3, column=0, columnspan=4, padx=10, pady=10, sticky="w")

    def clear_template_path(self):
        self.template_path.set("")
        self.status_message.set("Template file cleared. Please select a new file.")

    def clear_form_path(self):
        self.form_path.set("")
        self.status_message.set("Form file cleared. Please select a new file.")

    def browse_template(self):
        filename = filedialog.askopenfilename(title="Select Template File", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if filename:
            self.template_path.set(filename)
            self.status_message.set("Template file selected. Ready.")

    def browse_form(self):
        filename = filedialog.askopenfilename(title="Select Filled-in Form File", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if filename:
            self.form_path.set(filename)
            self.status_message.set("Form file selected. Ready.")

    # --- REVISED: Main workflow with file renaming ---
    def run_transfer(self):
        template_file = self.template_path.get()
        form_file = self.form_path.get()

        if not template_file or not form_file:
            messagebox.showwarning("Missing Files", "Please select both template and form files.")
            return
        
        directory = os.path.dirname(template_file)
        # Create a temporary name for the intermediate file
        temp_output_path = os.path.join(directory, f"temp_{os.path.basename(template_file)}")

        # Step 1: Transfer data to the temporary file
        self.status_message.set("Step 1/3: Transferring data...")
        self.master.update_idletasks()
        success, message = transfer_data_to_template(form_file, template_file, temp_output_path)

        if not success:
            self.status_message.set(message)
            messagebox.showerror("Error", message)
            if os.path.exists(temp_output_path): os.remove(temp_output_path)
            return

        # Step 2: Clean the temporary file and get the dossier name
        self.status_message.set("Step 2/3: Cleaning sheets...")
        self.master.update_idletasks()
        proc_success, proc_message, dossier_name = process_and_clean_sheets(temp_output_path)

        if not proc_success:
            self.status_message.set(proc_message)
            messagebox.showerror("Processing Error", proc_message)
            if os.path.exists(temp_output_path): os.remove(temp_output_path)
            return

        # Step 3: Rename the file to its final name
        self.status_message.set("Step 3/3: Finalizing and renaming file...")
        self.master.update_idletasks()
        try:
            clean_dossier_name = sanitize_filename(dossier_name)
            final_name = f"Aflossingstabel - {clean_dossier_name}.xlsx"
            final_path = os.path.join(directory, final_name)

            # Handle cases where the file already exists
            counter = 1
            name_part, ext_part = os.path.splitext(final_path)
            while os.path.exists(final_path):
                final_path = f"{name_part} ({counter}){ext_part}"
                counter += 1

            os.rename(temp_output_path, final_path)
            
            final_message = f"{message}\n{proc_message}\n\nFile saved as:\n{os.path.basename(final_path)}"
            self.status_message.set(final_message)
            messagebox.showinfo("Success", final_message)

        except Exception as e:
            final_message = f"Could not rename the final file. It has been saved as {os.path.basename(temp_output_path)}.\nError: {e}"
            self.status_message.set(final_message)
            messagebox.showerror("Rename Error", final_message)


# --- Run the GUI ---
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelTransferApp(root)
    root.mainloop()